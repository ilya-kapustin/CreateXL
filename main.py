from openpyxl.utils import rows_from_range
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.page import PageMargins


class StyleCell:

    def __init__(self, ws, cell, text='', border=False, border_sides='ALL', border_style='thin', border_color='000000',
                 horizontal='left', new_line=False, font_size=11, style_test='Times New Roman', bold_text=False,
                 merge_cells=False):
        self.ws = ws
        self.cell = cell
        self.text = text
        self.border = border
        self.border_sides = border_sides
        self.line_style = self.side(border_style, border_color)
        self.horizontal = horizontal
        self.new_line = new_line
        self.font_size = font_size
        self.style_test = style_test
        self.bold_text = bold_text
        self.merge_cells = merge_cells

    def run(self):
        if ':' in self.cell:
            self.many_cells(many_cells=self.cell, text=self.text, border=self.border, new_line=self.new_line,
                            horizontal=self.horizontal, bold_text=self.bold_text, border_sides=self.border_sides,
                            font_size=self.font_size, style_test=self.style_test, line_style=self.line_style)
            if self.merge_cells:
                self.merge(self.cell)
        else:
            self.one_cell(self.cell, self.text, self.new_line, self.horizontal, self.font_size, self.style_test,
                          self.bold_text)
            if self.border:
                self.border_cell(self.cell, self.border_sides, self.line_style)

    def merge(self, cell):
        self.ws.merge_cells(cell)

    def side(self, border_style, border_color):
        return Side(style=border_style, color=border_color)

    def one_cell(self, cell, text, new_line, horizontal, font_size, style_test, bold_text):
        self.settings(cell, new_line, horizontal, font_size, style_test, bold_text)
        self.ws[cell].value = text

    def many_cells(self, many_cells, text, border, new_line, horizontal, bold_text, border_sides, line_style, font_size,
                   style_test, marge=False):
        for cells in rows_from_range(many_cells):
            for cell in cells:
                self.settings(cell=cell, wrap_text=new_line, horizontal=horizontal, font_size=font_size,
                              style_test=style_test, bold_text=bold_text)
                if border:
                    self.border_cell(cell, border_sides, line_style)
                self.ws[cell].value = text
        if marge:
            self.merge(many_cells)

    def settings(self, cell, wrap_text, horizontal, font_size, style_test, bold_text):
        self.ws[cell].alignment = Alignment(horizontal=horizontal, wrapText=wrap_text)
        self.ws[cell].font = Font(size=font_size, name=style_test, bold=bold_text)

    def border_cell(self, cell, border_sides, line_style):
        if border_sides == 'ALL':
            self.ws[cell].border = Border(left=line_style, top=line_style, right=line_style,
                                          bottom=line_style)
        elif border_sides == 'left':
            self.ws[cell].border = Border(left=line_style)
        elif border_sides == 'top':
            self.ws[cell].border = Border(top=line_style)
        elif border_sides == 'right':
            self.ws[cell].border = Border(right=line_style)
        elif border_sides == 'bottom':
            self.ws[cell].border = Border(bottom=line_style)


class CreateFileXL:

    def __init__(self, save_name, paper_size='PAPERSIZE_A4', paper_orientation='ORIENTATION_PORTRAIT', cm=1 / 2.54):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.cm = cm
        self.save_name = save_name
        self.paperSize = paper_size
        self.paper_orientation = paper_orientation

    def create(self):
        self.settings_create()
        self.save_xl()

    def settings_create(self):
        self.ws.page_setup.paperSize = self.paper_settings()[0]
        self.ws.page_setup.orientation = self.paper_settings()[1]

    def paper_settings(self):
        paper_size = {
            'PAPERSIZE_LETTER': '1',
            'PAPERSIZE_LETTER_SMALL': '2',
            'PAPERSIZE_TABLOID': '3',
            'PAPERSIZE_LEDGER': '4',
            'PAPERSIZE_LEGAL': '5',
            'PAPERSIZE_STATEMENT': '6',
            'PAPERSIZE_EXECUTIVE': '7',
            'PAPERSIZE_A3': '8',
            'PAPERSIZE_A4': '9',
            'PAPERSIZE_A4_SMALL': '10',
            'PAPERSIZE_A5': '11',
        }
        paper_orientation = {
            'ORIENTATION_PORTRAIT': 'portrait',
            'ORIENTATION_LANDSCAPE': 'landscape',
        }
        return paper_size[self.paperSize], paper_orientation[self.paper_orientation]

    def save_xl(self):
        self.ws.page_margins = PageMargins(left=self.cm, right=self.cm, top=self.cm, bottom=self.cm)
        self.wb.save(self.save_name)


if __name__ == '__main__':
    obj = CreateFileXL(save_name='test123.xlsx', paper_orientation='ORIENTATION_LANDSCAPE', paper_size='PAPERSIZE_A5')
    obj.create()
    wb = load_workbook('test123.xlsx')
    ws = wb.active
    StyleCell(ws=ws, text='ergerg445ger45g45er54g4ergeg', cell='B2:G6', border=True, new_line=True,
              merge_cells=True).run()
    wb.save('test123.xlsx')
